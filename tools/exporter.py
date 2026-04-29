#!/usr/bin/env python3
"""
Keyword Gap Export Module
==========================
Accepts a filtered pandas DataFrame and exports it to .tmp/ as CSV and/or
XLSX. Filename convention: keyword_gap_legendary_parts_{competitor}_{date}.{ext}

Reuses the autosize_columns pattern from tools/semrush_competitor_keywords.py
for consistent Excel formatting across all project exports.

Dependencies:
    pip install pandas openpyxl

Usage:
    from tools.exporter import export
    paths = export(df, competitor_domain="partseurope.eu")
    # returns {"csv": "/path/to/.tmp/keyword_gap_...csv",
    #          "xlsx": "/path/to/.tmp/keyword_gap_...xlsx"}
"""

import os
import re
import logging
from typing import Optional
import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment

log = logging.getLogger(__name__)

OUTPUT_DIR = ".tmp"
HEADER_FILL_COLOR = "BDD7EE"   # Light blue, matches existing Excel exports
TODO_FILL_COLOR   = "FFFF00"   # Yellow, flags cells that need manual follow-up


def build_filename(competitor_domain: str, date: str, extension: str) -> str:
    """
    Build the output filename following project naming convention.

    Sanitises competitor_domain by replacing dots with underscores and
    stripping characters outside [a-zA-Z0-9_-] to produce a safe filename.

    Args:
        competitor_domain: Raw domain string (e.g. "partseurope.eu").
        date:              Date string in YYYYMMDD format.
        extension:         File extension without dot ("csv" or "xlsx").

    Returns:
        Filename string (not a full path).

    Example:
        build_filename("partseurope.eu", "20260413", "csv")
        → "keyword_gap_legendary_parts_partseurope_eu_20260413.csv"
    """
    safe_domain = competitor_domain.replace(".", "_")
    safe_domain = re.sub(r"[^a-zA-Z0-9_\-]", "", safe_domain)
    return f"keyword_gap_legendary_parts_{safe_domain}_{date}.{extension}"


def build_multi_filename(competitor_domains: list, date: str, extension: str) -> str:
    """
    Build the output filename for multi-competitor exports.

    Joins up to 2 domain names with '_vs_' to keep filenames readable.
    If more than 2 competitors, only the first two are used and
    '_and_more' is appended.

    Args:
        competitor_domains: List of raw domain strings.
        date:               Date string in YYYYMMDD format.
        extension:          File extension without dot ("csv" or "xlsx").

    Returns:
        Filename string (not a full path).

    Example:
        build_multi_filename(["autozone.com", "partseurope.eu"], "20260415", "xlsx")
        → "keyword_gap_legendary_parts_autozone_com_vs_partseurope_eu_20260415.xlsx"
    """
    def safe(domain: str) -> str:
        s = domain.replace(".", "_")
        return re.sub(r"[^a-zA-Z0-9_\-]", "", s)

    if len(competitor_domains) == 1:
        return build_filename(competitor_domains[0], date, extension)

    parts = [safe(d) for d in competitor_domains[:2]]
    suffix = "_and_more" if len(competitor_domains) > 2 else ""
    joined = "_vs_".join(parts) + suffix
    return f"keyword_gap_legendary_parts_{joined}_{date}.{extension}"


def build_backlink_filename(competitor_domain: str, date: str = None, extension: str = "xlsx") -> str:
    """
    Build the output filename for backlink gap exports.

    Sanitises competitor_domain by replacing dots and special chars with underscores.

    Args:
        competitor_domain: Raw domain string (e.g. "partseurope.eu").
        date:              Date string in YYYYMMDD format. Defaults to today.
        extension:         File extension without dot ("csv" or "xlsx").

    Returns:
        Filename string (not a full path).

    Example:
        build_backlink_filename("partseurope.eu")
        → "backlink_gap_partseurope_eu_20260414.xlsx"
    """
    if date is None:
        date = datetime.now().strftime("%Y%m%d")
    safe_domain = competitor_domain.replace(".", "_")
    safe_domain = re.sub(r"[^a-zA-Z0-9_\-]", "", safe_domain)
    return f"backlink_gap_{safe_domain}_{date}.{extension}"


def autosize_columns(ws) -> None:
    """
    Widen each Excel column to fit its longest cell value (capped at 65 chars).

    Ported verbatim from tools/semrush_competitor_keywords.py to maintain
    consistent Excel formatting across all project exports.

    Args:
        ws: openpyxl Worksheet object.
    """
    for col_cells in ws.columns:
        width = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(width + 3, 65)


def style_header_row(ws) -> None:
    """
    Apply bold font, light-blue fill, and center alignment to the header row.

    Applied to row 1 (the column header row written by DataFrame.to_excel).

    Args:
        ws: openpyxl Worksheet object (after DataFrame has been written).
    """
    fill = PatternFill(start_color=HEADER_FILL_COLOR, end_color=HEADER_FILL_COLOR,
                       fill_type="solid")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = bold
        cell.alignment = center


def highlight_todo_cells(ws) -> None:
    """
    Apply a yellow fill to every cell in the "Email" column whose value is "TODO".

    Called from export_xlsx() after style_header_row() so the header styling
    is not overwritten. Does nothing if no "Email" column exists.

    Args:
        ws: openpyxl Worksheet object (after DataFrame has been written).
    """
    # Find the column index of "Email" from the header row
    email_col = None
    for cell in ws[1]:
        if str(cell.value).strip().lower() == "email":
            email_col = cell.column
            break

    if email_col is None:
        return

    yellow = PatternFill(start_color=TODO_FILL_COLOR, end_color=TODO_FILL_COLOR,
                         fill_type="solid")
    for row in ws.iter_rows(min_row=2, min_col=email_col, max_col=email_col):
        for cell in row:
            if cell.value == "TODO":
                cell.fill = yellow


def export_csv(df: pd.DataFrame, filepath: str) -> None:
    """
    Write DataFrame to a UTF-8 CSV file with header row.

    Args:
        df:       Sorted DataFrame ready for export.
        filepath: Absolute or relative path for the output file.
    """
    df.to_csv(filepath, index=False, encoding="utf-8")
    log.info("CSV written → %s (%d rows)", filepath, len(df))


def export_xlsx(df: pd.DataFrame, filepath: str, sheet_name: str = "Keyword Gap") -> None:
    """
    Write DataFrame to an xlsx file using openpyxl with styled headers
    and auto-sized columns.

    Uses pd.ExcelWriter with engine="openpyxl" — same pattern as
    semrush_competitor_keywords.py. Applies style_header_row() and
    autosize_columns() after writing.

    Args:
        df:         Sorted DataFrame ready for export.
        filepath:   Absolute or relative path for the output file.
        sheet_name: Excel sheet tab name. Defaults to "Keyword Gap".
    """
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        style_header_row(ws)
        highlight_todo_cells(ws)
        autosize_columns(ws)
    log.info("XLSX written → %s (%d rows)", filepath, len(df))


def export(
    df: pd.DataFrame,
    competitor_domain,
    date: Optional[str] = None,
    output_dir: str = OUTPUT_DIR,
    formats: Optional[list] = None,
) -> dict:
    """
    Main export entry point. Sorts, validates, and writes to all requested formats.

    Sort order: Volume descending (primary), KD ascending (secondary tiebreak).
    Creates output_dir if it does not exist.

    Args:
        df:               Filtered DataFrame from agent_workflow.py.
        competitor_domain: Single domain string OR list of domain strings.
                           A list triggers multi-competitor filename format.
        date:             Date string YYYYMMDD. Defaults to today.
        output_dir:       Target directory. Default ".tmp".
        formats:          List of "csv" and/or "xlsx". Defaults to both.

    Returns:
        Dict mapping format name to absolute output path.
        Example: {"csv": "/path/.tmp/keyword_gap_....csv",
                  "xlsx": "/path/.tmp/keyword_gap_....xlsx"}

    Raises:
        ValueError: If df is empty.
        OSError:    If output_dir cannot be created or written to.
    """
    if df.empty:
        raise ValueError("Cannot export: DataFrame is empty. No keywords matched the filters.")

    if date is None:
        date = datetime.now().strftime("%Y%m%d")

    if formats is None:
        formats = ["csv", "xlsx"]

    # Sort: Volume DESC, KD ASC as tiebreak
    sort_cols = []
    if "Volume" in df.columns:
        sort_cols.append(("Volume", False))
    if "KD" in df.columns:
        sort_cols.append(("KD", True))

    if sort_cols:
        cols, ascending = zip(*sort_cols)
        df = df.sort_values(list(cols), ascending=list(ascending)).reset_index(drop=True)

    os.makedirs(output_dir, exist_ok=True)

    # Choose filename builder based on whether one or multiple competitors
    use_multi = isinstance(competitor_domain, list)

    output_paths = {}
    for fmt in formats:
        if use_multi:
            filename = build_multi_filename(competitor_domain, date, fmt)
        else:
            filename = build_filename(competitor_domain, date, fmt)
        filepath = os.path.abspath(os.path.join(output_dir, filename))
        if fmt == "csv":
            export_csv(df, filepath)
        elif fmt == "xlsx":
            export_xlsx(df, filepath)
        else:
            log.warning("Unknown format '%s' — skipping", fmt)
            continue
        output_paths[fmt] = filepath

    return output_paths
